name = ' bob'  # 자료형을 따로 안만들어도 됨

num = 12

print (name, num)

# 리스트
a_list = []

a_list.append(1)

a_list.append([1, 2])
print(a_list)

# dic
a_dict = {}


# 함수
def sum_all(a, b, c):
    return a + b + c


#퀴즈
def name(name):
    if(name=="김민준"):
        print (name)



name("김민준")
name("김준")


#반복문
for_list =[1,2,3,4,5]

def sum(mylist):
    sum=0
    for i in mylist:
        sum =sum+i
    print(sum)


sum(for_list)


#range

rang_list =range(10)
print(rang_list[0])


#이름만뽑기
for_list



a = 'spartacodingclub@gmail.com'

#채워야하는 함수
def check_mail(s):   # 여기에 코딩을 해주세요  #결과값
    for i in a:
        if(i=='@'):
            print('true')



check_mail(a)

#아래와 같이 출력됩니다
True


#입력값
a = ['사과','감','감','배','포도','포도','딸기','포도','감','수박','딸기']

#채워야하는 함수
fruit = {}


def count_list(a_list): # 여기에 코딩을 해주세요   #결과값
 for i in a:
    print(i)
    if i in fruit:
        fruit[i]+=1
    else:
        fruit[i]=1



count_list(a)
print(fruit)




#아래와 같이 출력됩니다
{'사과': 1, '감': 3, '배': 1, '포도': 3, '딸기': 2, '수박': 1}


#requests
import requests # requests 라이브러리 설치 필요

r = requests.get('http://openapi.seoul.go.kr:8088/6d4d776b466c656533356a4b4b5872/json/RealtimeCityAir/1/99')
rjson = r.json()
print (rjson['RealtimeCityAir']['row'][0]['NO2'])


#크롤링 배우기
import requests
from bs4 import BeautifulSoup

# URL을 읽어서 HTML를 받아오고,
headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)

# HTML을 BeautifulSoup이라는 라이브러리를 활용해 검색하기 용이한 상태로 만듦
soup = BeautifulSoup(data.text, 'html.parser')

# select를 이용해서, tr들을 불러오기
movies = soup.select('#old_content > table > tbody > tr')

# movies (tr들) 의 반복문을 돌리기

for movie in movies:
    # movie 안에 a 가 있으면,
    a_tag = movie.select_one('td.title > div > a')
    a_ac = movie.select_one('td.point')

    if a_ac is not None:
        a_ac = movie.select_one('td.point').text
    if a_tag is not None:
        # a의 text를 찍어본다.
        print (a_tag.text, a_ac)


#엑셀 활용하기 openpyxl file -> setting 에서 install
from openpyxl import load_workbook

work_book = load_workbook('prac01.xlsx')
work_sheet = work_book['prac']

# 데이터를 읽어봅니다.
temp_text = work_sheet.cell(row = 1, column = 1).value # 1행 1열을 읽어온다
print (temp_text)
work_sheet.cell(row=3, column=2, value='홍길동')

# 수정한 엑셀파일을 저장합니다.
# 참고: 다른이름으로 저장할 수도 있습니다.
work_book.save('prac01.xlsx')


for movie in movies:
    # movie 안에 a 가 있으면,
    a_tag = movie.select_one('td.title > div > a')
    a_ac = movie.select_one('td.point')

    if a_ac is not None:
        a_ac = movie.select_one('td.point').text
    if a_tag is not None:
        # a의 text를 찍어본다.
        print (a_tag.text, a_ac)


#실습
    i=0
    for movie in movies:
        a_tag = movie.select_one('td.title > div > a')
        a_ac = movie.select_one('td.point')

        if a_tag is not None:
             i = i + 1
             work_sheet.cell(row=i, column=1, value=i)
             work_sheet.cell(row=i, column=2, value=a_tag.text)
             work_book.save('prac01.xlsx')

        if a_ac is not None:
                work_sheet.cell(row=i, column=3, value=a_ac.text)
                work_book.save('prac01.xlsx')