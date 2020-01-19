
#app.run('127.0.0.1',port=5000,debug=True)

from flask import Flask , render_template , jsonify ,request, session, redirect, url_for  # render_template 안에서 html 사용
import requests
from bs4 import BeautifulSoup

#엑셀 파일 생성 시 필요
from openpyxl import load_workbook, Workbook

import jwt
import datetime   #로그인시 토큰시간 지정
import hashlib     #비밀번호 해쉬 암호화 db 저장





app = Flask(__name__)

from pymongo import MongoClient           # pymongo를 임포트 하기(패키지 인스톨 먼저 해야겠죠?)

client = MongoClient('localhost', 27017)  # mongoDB는 27017 포트로 돌아갑니다.
db = client.dbsparta                      # 'dbsparta'라는 이름의 db를 만듭니다.


## HTML을 주는 부분

SECRET_KEY = 'apple'  # 토근시 필요한 보안 - 아무거나 입력 원하는거

@app.route('/register')         # 회원이 아니라 면을 클릭하게 되면
def register():
   return render_template('register.html')


@app.route('/api/register', methods=['POST'])  #회원가입시     db.counting.insert_one(imformation)
def api_register():
   id_receive = request.form['id_give']
   pw_receive = request.form['pw_give']


   pw_hash = hashlib.sha256(pw_receive.encode('utf-8')).hexdigest()

   db.temp.insert_one({'id':id_receive,'pw':pw_hash})

   return jsonify({'result': 'success'})


@app.route('/api/login', methods=['POST'])
#로그인 완료시  id와 pwd 비교 성공하면 html 에서  index.html로 이동
def api_login():
   id_receive = request.form['id_give']
   pw_receive = request.form['pw_give']

   # 회원가입 때와 같은 방법으로 pw를 암호화합니다.
   pw_hash = hashlib.sha256(pw_receive.encode('utf-8')).hexdigest()

   # id, 암호화된pw을 가지고 해당 유저를 찾습니다.
   result = db.temp.find_one({'id':id_receive,'pw':pw_hash})

   # 찾으면 JWT 토큰을 만들어 발급합니다.
   if result is not None:
      # JWT 토큰에는, payload와 시크릿키가 필요합니다.
      # 시크릿키가 있어야 토큰을 디코딩(=풀기) 해서 payload 값을 볼 수 있습니다.
      # 아래에선 id와 exp를 담았습니다. 즉, JWT 토큰을 풀면 유저ID 값을 알 수 있습니다.
      # exp에는 만료시간을 넣어줍니다. 만료시간이 지나면, 시크릿키로 토큰을 풀 때 만료되었다고 에러가 납니다.
      payload = {
         'id': id_receive,
         'exp': datetime.datetime.utcnow() + datetime.timedelta(seconds=3600)
      }
      token = jwt.encode(payload, SECRET_KEY, algorithm='HS256').decode('utf-8')

      # token을 줍니다.
      return jsonify({'result': 'success','token':token})
   # 찾지 못하면
   else:
      return jsonify({'result': 'fail', 'msg':'아이디/비밀번호가 일치하지 않습니다.'})


@app.route('/api/id', methods=['GET'])
def api_valid():
   # 토큰을 주고 받을 때는, 주로 header에 저장해서 넘겨주는 경우가 많습니다.
   # header로 넘겨주는 경우, 아래와 같이 받을 수 있습니다.
   token_receive = request.headers['token_give']

   # try / catch 문?
   # try 아래를 실행했다가, 에러가 있으면 except 구분으로 가란 얘기입니다.

   try:
      # token을 시크릿키로 디코딩합니다.
      # 보실 수 있도록 payload를 print 해두었습니다. 우리가 로그인 시 넣은 그 payload와 같은 것이 나옵니다.
      payload = jwt.decode(token_receive, SECRET_KEY, algorithms=['HS256'])
     # print(payload)

      # payload 안에 id가 들어있습니다. 이 id로 유저정보를 찾습니다.
      # 여기에선 그 예로 닉네임을 보내주겠습니다.
      userinfo = db.temp.find_one({'id':payload['id']},{'_id':0})
      return jsonify({'result': 'success','id':userinfo['id']})
   except jwt.ExpiredSignatureError:
      # 위를 실행했는데 만료시간이 지났으면 에러가 납니다.
      return jsonify({'result': 'fail', 'msg':'로그인 시간이 만료되었습니다.'})


@app.route('/')
def home():
   return render_template('index2.html')


@app.route('/index2.html')
def index2():
   return render_template('index2.html')

@app.route('/index.html')
def index():
   return render_template('index.html')

@app.route('/input.html')
def input():
   return render_template('input.html')

@app.route('/excel.html')
def excel():
   return render_template('excel.html')

@app.route('/entire.html')
def entire():
   return render_template('entire.html')


@app.route('/month.html')
def month():
   return render_template('month.html')




## API 역할을 하는 부분
@app.route('/money', methods=['POST'])
def saving():
   id =request.form['id_give']
   date = request.form['someDate_give']
   pay = request.form['money_give']
   cont = request.form['content_give']
   tag = request.form['tag_give']

   # 클라이언트로부터 데이터를 받는 부분

   # mongoDB에 넣는 부분

   imformation = {
      'id': id,
      'somedate': date,
      'money':pay,
      'content':cont,
      'tag':tag,

   }

   db.counting.insert_one(imformation)


   return jsonify({'result': 'success'})




@app.route('/money', methods=['GET'])
def getting():
   # 모든 document 찾기 & _id 값은 출력에서 제외하기
  # result =list(db.counting.find({},{'_id':0}))
   result =list(db.counting.find({},{'_id':0}))
   # shops라는 키 값으로  내려주기
   return jsonify({'result': 'success', 'shops': result})

#엑셀 생성시
@app.route('/excel',methods=['GET'])
def exeting():
   result = list(db.counting.find({}, {'_id': 0}))
   row = 4
   token_receive = request.headers['token_give']



   #payload를 찍어보면 id 값이 존재한다
   payload = jwt.decode(token_receive, SECRET_KEY, algorithms=['HS256'])

      # 해당 아이디 값을 받아서 엑셀 저장 이름에 넣는다.
   id = payload['id'] + '.xlsx'
   work_book = Workbook()


   print(id)
   work_sheet = work_book.active
   work_sheet.title="par"
   sum=0
   cafe=0
   food=0
   medical=0
   other=0
   work_sheet.cell(row=3, column=1, value="날짜")  # 날짜 엑셀에 적용
   work_sheet.cell(row=3, column=2, value="내용")  # 날짜 엑셀에 적용
   work_sheet.cell(row=3, column=3, value="종류")  # 날짜 엑셀에 적용
   work_sheet.cell(row=3, column=4, value="금액")  # 날짜 엑셀에 적용

   work_sheet.cell(row=3, column=6, value="총금액")  # 날짜 엑셀에 적용

   work_sheet.cell(row=6, column=6, value="카페")  # 날짜 엑셀에 적용
   work_sheet.cell(row=6, column=7, value="의료")  # 날짜 엑셀에 적용
   work_sheet.cell(row=6, column=8, value="음식")  # 날짜 엑셀에 적용
   work_sheet.cell(row=6, column=9, value="기타")  # 날짜 엑셀에 적용
   for s in result:
   # print(s['somedate'])
    if(payload['id']==s['id']):
     work_sheet.cell(row=row, column=1, value=s['somedate']) #날짜 엑셀에 적용
     work_sheet.cell(row=row, column=2, value=s['content'])  # 내용 엑셀에 적용
     work_sheet.cell(row=row, column=3, value=s['tag'])  # 태그 엑셀에 적용
     work_sheet.cell(row=row, column=4, value=s['money'])  #돈  엑셀에 적용
     sum+=int(s['money'])
     if(s['tag']=='카페'):
       cafe+=int(s['money'])
     if (s['tag'] == '카페'):
       cafe += int(s['money'])
     if (s['tag'] == '음식'):
       food += int(s['money'])
     if (s['tag'] == '의료비 및 보험비'):
        medical += int(s['money'])
     if (s['tag'] == '기타'):
        other += int(s['money'])
     row=row+1

   work_sheet.cell(row=4, column=6, value=sum)  # 총 금액 엑셀에 적용

    #태그별 금액
   work_sheet.cell(row=7, column=6, value=cafe)
   work_sheet.cell(row=7, column=7, value=food)
   work_sheet.cell(row=7, column=8, value=medical)
   work_sheet.cell(row=7, column=9, value=other)
   work_book.save(id)



   return jsonify({'result': 'success', 'shops': result})


if __name__ == '__main__':
   app.run('127.0.0.1',port=5000,debug=True)
   #aws 0.0.0.0
   #127.0.0.1